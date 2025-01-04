import pytest
import csv
import json
from openpyxl import load_workbook
from src.utils.export_helpers import *



@pytest.fixture
def sample_expenses():
    return [
        {"ID": 1, "Date": "2024-01-01", "Amount": 50.0, "Category": "Groceries", "Description": "Groceries shopping"},
        {"ID": 2, "Date": "2024-01-05", "Amount": 20.0, "Category": "Leisure", "Description": "Movie night"},
    ]

@pytest.fixture
def sample_budget_info():
    return {"budget_amount": 1000.0, "current_expenses": 70.0, "remaining_budget": 930.0}



def test_write_csv(tmp_path, sample_expenses, sample_budget_info):
    output_path = tmp_path / "expenses.csv"
    write_csv(output_path, sample_expenses, sample_budget_info)

    with open(output_path, "r", encoding="utf-8") as file:
        reader = csv.reader(file)
        rows = list(reader)

    assert rows[0] == ["ID", "Date", "Amount", "Category", "Description"]
    assert rows[1] == ["1", "2024-01-01", "50.0", "Groceries", "Groceries shopping"]
    assert rows[2] == ["2", "2024-01-05", "20.0", "Leisure", "Movie night"]

    assert rows[3] == ["", "", "", "", ""]
    assert rows[4] == ["Budget Information", "", "", "", ""]
    assert rows[5][1:] == ["Budget Amount", "$1000.00", "", ""]
    assert rows[6][1:] == ["Current Expenses", "$70.00", "", ""]
    assert rows[7][1:] == ["Remaining Budget", "$930.00", "", ""]


def test_write_json(tmp_path, sample_expenses, sample_budget_info):
    output_path = tmp_path / "expenses.json"
    write_json(output_path, sample_expenses, sample_budget_info)

    with open(output_path, "r", encoding="utf-8") as file:
        content = json.load(file)

    assert "expenses" in content
    assert len(content["expenses"]) == 2
    assert content["expenses"][0]["ID"] == 1
    assert content["expenses"][0]["Date"] == "2024-01-01"

    assert "Budget Information" in content
    assert content["Budget Information"]["Budget Amount"] == "$1000.00"
    assert content["Budget Information"]["Remaining Budget"] == "$930.00"


def test_write_excel(tmp_path, sample_expenses, sample_budget_info):
    output_path = tmp_path / "expenses.xlsx"
    write_excel(output_path, sample_expenses, sample_budget_info)

    wb = load_workbook(output_path)
    ws = wb["Expenses"]

    headers = [cell.value for cell in ws[1]]
    assert headers == ["ID", "Date", "Amount", "Category", "Description"]

    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "2024-01-01"
    assert ws.cell(row=2, column=3).value == 50.0

    assert ws.cell(row=5, column=1).value == "Budget Information"
    assert ws.cell(row=6, column=1).value == "Budget Amount"
    assert ws.cell(row=6, column=2).value == "$1000.00"



def test_generate_unique_filename(tmp_path):
    existing_file = tmp_path / "file.json"
    existing_file.touch()

    unique_file = generate_unique_filename(existing_file)
    assert unique_file.name == "file(1).json"

    unique_file.touch()
    next_unique_file = generate_unique_filename(existing_file)
    assert next_unique_file.name == "file(2).json"



def test_filter_expenses(sample_expenses):
    filtered = filter_expenses(sample_expenses, year=2024)
    assert len(filtered) == 2

    filtered = filter_expenses(sample_expenses, month=1)
    assert len(filtered) == 2

    filtered = filter_expenses(sample_expenses, category="Groceries")
    assert len(filtered) == 1
    assert filtered[0]["Category"] == "Groceries"
